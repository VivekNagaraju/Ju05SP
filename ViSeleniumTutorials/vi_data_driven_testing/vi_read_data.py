'''
Created on 17-Sep-2025

@author: admin
'''
import openpyxl

# 1. Load the workbook 
filename = r"C:\Users\admin\Desktop\vivek.xlsx"
my_workbook = openpyxl.load_workbook(filename)

# 2. Get the sheet
# active_sheet = my_workbook.active # to get the active
active_sheet = my_workbook["Sheet1"] # to get the sheet by name

# 3. Get the no. of rows and columns
total_rows = active_sheet.max_row
print(total_rows)

total_cols = active_sheet.max_column
print(total_cols)

# 4. Get the data
for i in range(2, total_rows+1):
    username = active_sheet.cell(i, 1).value
    password = active_sheet.cell(i, 2).value
    print(username, password)